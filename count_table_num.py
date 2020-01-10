from openpyxl import load_workbook
import warnings
import csv

warnings.filterwarnings("ignore")

#获取表总数
def get_table_name(excel,sheet):
    table = excel.get_sheet_by_name(sheet)
    rows = table.max_row - 2
    return rows

#获取所有表字段总数
def get_field_name(excel,sheets):
    sum = 0
    for sheet in sheets:
        table = excel.get_sheet_by_name(sheet)
        rows = table.max_row - 2
        sum += rows
    return sum

#根据字段获取字段明细（列表）
def get_XB_frequency(excel,sheets):
    XB = []
    for i in range(0, len(sheets)):
        table = excel.get_sheet_by_name(sheets[i])
        max_row = table.max_row
        for row in table.iter_rows(min_row=3, min_col=2 ,max_col=2, max_row=max_row):
            for cell in row:
                if cell.value == 'XB':
                    XB.append(table.title)
    return len(XB),XB

#获取所有表的所有字段（列表）
def get_all_field(excel,sheets):
    all_field = []
    for i in range(0,len(sheets)):
        table = excel.get_sheet_by_name(sheets[i])
        max_row = table.max_row
        for row in table.iter_rows(min_row=3,max_row=max_row,min_col=2,max_col=2):
            for cell in row:
                all_field.append(cell.value)
    return all_field

#获取单个说明的字段明细（字典）
def get_one_field_description(excel,sheets,field):
    count = 1
    data = ''
    field_description = {}
    for i in range(0, len(sheets)):
        table = excel.get_sheet_by_name(sheets[i])
        max_row = table.max_row
        for row in table.iter_rows(min_row=3, max_row=max_row, min_col=2, max_col=6):
            for cell in row:
                if count % 5 == 1:
                    data = cell.value
                if count % 5 == 0 and field in str(cell.value):
                    if data in field_description.keys():
                        field_description[data] += 1
                    else:
                        field_description[data] = 1
                count += 1
    return field_description

#获取所有说明的明细（字典）
def get_description_detail(excel,sheets):
    all_field_description = {}
    for i in range(0, len(sheets)):
        table = excel.get_sheet_by_name(sheets[i])
        max_row = table.max_row
        for row in table.iter_rows(min_row=3, max_row=max_row, min_col=6, max_col=6):
            for cell in row:
                if str(cell.value) in all_field_description.keys() and str(cell.value) != '':
                    all_field_description[cell.value] += 1
                else:
                    all_field_description[cell.value] = 1
    del all_field_description[None]
    return all_field_description

#获取去重后的字段（列表）
def get_list_set_field(excel,sheets):
    all_field = []
    for i in range(0, len(sheets)):
        table = excel.get_sheet_by_name(sheets[i])
        max_row = table.max_row
        for row in table.iter_rows(min_row=3, max_row=max_row, min_col=2, max_col=2):
            for cell in row:
                all_field.append(cell.value)
    return list(set(all_field))

#获取去重后的字段（列表）
def get_list_set_description(excel,sheets):
    all_field = []
    for i in range(0, len(sheets)):
        table = excel.get_sheet_by_name(sheets[i])
        max_row = table.max_row
        for row in table.iter_rows(min_row=3, max_row=max_row, min_col=6, max_col=6):
            for cell in row:
                all_field.append(cell.value)
    return list(set(all_field))

def get_field_num(all_field,x):
    return all_field.count(x)

def dict_sort(d):
    return dict(sorted(d.items(),key=lambda d:d[1],reverse=True))

def write_to_csv(result):
    with open('result.csv','a',newline='',encoding='gb18030') as csvfile:
        filenames = ['去重字段名称','数量','明细']
        write = csv.DictWriter(csvfile,fieldnames=filenames)
        # write.writeheader()
        result_dict = dict(map(lambda x,y:[x,y],filenames,result))
        result_dict['明细'] = ','.join(result[2])
        print(result_dict)
        write.writerow(result_dict)

if __name__ == '__main__':
    excel = load_workbook('E:\公司资料\大数据实验室\xxx/芜湖表结构.xlsx')
    sheets = excel.get_sheet_names()

    all_description_description = get_description_detail(excel,sheets[1:])
    all_description_description_sort = dict_sort(all_description_description)
    print(all_description_description_sort)

    num = 0
    for v in all_description_description_sort.values():
        num += v

    for k in all_description_description_sort.keys():
        result = []
        result.append(k)
        result.append(all_description_description_sort[k])
        field_description = get_one_field_description(excel,sheets[1:],k)
        field_description_list = list(field_description)
        result.append(field_description_list)
        write_to_csv(result)

    # #获取表总数1459
    # table_num = len(sheets) - 1
    # print("表总数：%s" % table_num)
    #
    # #获取所有表的字段总数32312
    # field_num = get_field_name(excel,sheets[1:])
    # print("字段总数：%s" % field_num)
    #
    # #获取去重后的字段总数8977
    # field_set_num = len(get_list_set_description(excel,sheets))
    # print("去重后的字段总数：%s" % field_set_num)

    # all_set_description = get_list_set_description(excel,sheets[1:])
    # print(all_set_description)

    # all_field_description = get_description_detail(excel,sheets)
    # all_field_description = dict_sort(all_field_description)
    # print(dict_sort(all_field_description))

    # for k in all_field_description.keys():
    #     field_description = get_one_field_description(excel,sheets,k)
    #     print(field_description)

    # length,XB = get_XB_frequency(excel,sheets[1:])
    # all_field = get_all_field(excel,sheets[1:])

    # xb_all_description = get_all_xb_description(excel,sheets[1:])
    # print(xb_all_description)

    # xb_num = get_field_num(all_field, 'XM')
    # print("姓名的频率为：%s" % xb_num)
    #
    # xb_num = get_field_num(all_field, 'MZ')
    # print("名族的频率为：%s" % xb_num)
    #
    # xb_num = get_field_num(all_field, 'JG')
    # print("籍贯的频率为：%s" % xb_num)
    #
    # xb_num = get_field_num(all_field, 'XB')
    # print("性别的频率为：%s" % xb_num)
    #
    # xb_num = get_field_num(all_field, 'CSRQ')
    # print("出生日期的频率为：%s" % xb_num)
    #
    # xb_num = get_field_num(all_field, 'RKGLLB')
    # print("人口管理类别的频率为：%s" % xb_num)
    #
    # xb_num = get_field_num(all_field, 'GMSFHM')
    # print("公民身份号码的频率为：%s" % xb_num)
